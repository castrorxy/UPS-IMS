import datetime
import io
import json
import os
import sqlite3
import sys
import threading
import webbrowser
from functools import wraps

from flask import Flask, g, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook, load_workbook
from io import BytesIO
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash
import traceback

if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(APP_DIR)
DATA_DIR = os.path.join(ROOT_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "inventory.db")

FIELDS = [
    "stock_no",
    "code",
    "category",
    "system",
    "model",
    "rating",
    "status",
    "serial_number",
    "client",
    "date_acquired",
    "date_installed",
    "dr_no",
    "si_no",
    "po",
    "value_vat_ex",
    "warranty",
    "terms",
    "remarks",
]

DEFAULT_OPTIONS = {
    "code": [
        "UPS-KOT-10KVA/220V",
        "UPS-KOT-1KVA/220V",
        "UPS-KOT-2KVA/220V",
        "UPS-KOT-3KVA/220V",
    ],
    "brand": [
        "KOTOHIRA SYSTEM",
        "TM1110S",
        "PROLINK",
        "ABLEREX",
    ],
    "model": [
        "BH20S-V-RM",
        "BH60S",
        "BH30S-V",
        "BHT 1KS",
        "TM1110S",
        "6KVAS",
        "TGD11",
        "GH11",
        "PROLINK",
        "BH10S-V",
        "TGD12",
        "TGD13",
        "TGD14",
        "TGD15",
        "TGD16",
        "TGD17",
        "TGD18",
        "TGD19",
        "TGD20",
        "TGD21",
        "TGD22",
        "TGD23",
        "TGD24",
        "TGD25",
        "TGD26",
        "TGD27",
        "TGD28",
        "TGD29",
        "TGD30",
        "ARES PLUS RT3K-3KVA/2700 MA8060F8P0084",
        "BH100S-V",
    ],
    "rating": [
        "2KVA",
        "6KVA/6KW",
        "3KVA/3KW",
        "1KVA",
        "10KVA/10KW",
    ],
    "status": [
        "ON STOCK",
        "DELIVERED",
    ],
    "client": [
        "BE3 POWER SOLUTIONS ENTERPRISE",
        "GUILLBERN CORP.",
        "LRES",
        "SAN BEDA/SYNERFIVE",
        "SANSYU P. LIPA",
        "TERRAFEED VENTURES INC.",
        "TOUCHSTAR MEDICAL ENTERPRISES INC.",
        "ACQUISITION SERVICE IMPLEMENTATION CONSTRUCTION INC.",
        "DATANET",
    ],
    "remarks": [
        "VAT INC.",
        "Office",
        "RETURN FOR CHECKING UNIT",
        "Las Piñas",
        "Repaired",
    ],
}

# Add category choices for add-item forms
DEFAULT_OPTIONS["category"] = ["UPS", "AVR"]

CATEGORY_META = [
    {"key": "category", "label": "Category"},
    {"key": "code", "label": "Code"},
    {"key": "brand", "label": "Brand"},
    {"key": "remarks", "label": "Remarks"},
    {"key": "model", "label": "Model"},
    {"key": "rating", "label": "Rating"},
    {"key": "status", "label": "Status"},
    {"key": "client", "label": "Client"},
]

HEADER_MAP = {
    "STOCK #": "stock_no",
    "STOCK NO": "stock_no",
    "CODE": "code",
    "CATEGORY": "category",
    "COLUMN1": "system",
    "SYSTEM": "system",
    "MODEL": "model",
    "RATING": "rating",
    "STATUS": "status",
    "SERIAL NUMBER": "serial_number",
    "SERIAL NO": "serial_number",
    "CLIENT": "client",
    "DATE ACQUIRED": "date_acquired",
    "DATE INSTALLED": "date_installed",
    "DR NO.": "dr_no",
    "DR NO": "dr_no",
    "SI NO.": "si_no",
    "SI NO": "si_no",
    "PO": "po",
    "VALUE VAT EX": "value_vat_ex",
    "WARRANTY": "warranty",
    "TERMS": "terms",
    "REMARKS": "remarks",
}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")
_startup_done = False


@app.after_request
def add_no_cache_headers(response):
    if request.path.startswith("/api/") or request.path in {
        "/",
        "/inventory",
        "/clients",
        "/installations",
        "/warranty",
        "/reports",
        "/options",
    }:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
    return response


def get_db():
    if "db" not in g:
        os.makedirs(DATA_DIR, exist_ok=True)
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(error):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = get_db()
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT NOT NULL,
            actor TEXT NOT NULL,
            details TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            stock_no TEXT,
            code TEXT,
            category TEXT,
            system TEXT,
            model TEXT,
            rating TEXT,
            status TEXT,
            serial_number TEXT,
            client TEXT,
            date_acquired TEXT,
            date_installed TEXT,
            dr_no TEXT,
            si_no TEXT,
            po TEXT,
            value_vat_ex TEXT,
            warranty TEXT,
            terms TEXT,
            remarks TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS inventory_options (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            value TEXT NOT NULL,
            color TEXT,
            UNIQUE(category, value)
        )
        """
    )
    db.commit()

    # ensure color column exists for older databases
    cols = [r['name'] for r in db.execute("PRAGMA table_info(inventory_options)").fetchall()]
    if 'color' not in cols:
        try:
            db.execute("ALTER TABLE inventory_options ADD COLUMN color TEXT")
            db.commit()
        except Exception:
            db.rollback()


def ensure_users_schema():
    db = get_db()
    columns = [row["name"] for row in db.execute("PRAGMA table_info(users)").fetchall()]
    if "display_name" not in columns:
        db.execute("ALTER TABLE users ADD COLUMN display_name TEXT")
        db.commit()


def ensure_inventory_options():
    db = get_db()
    for category, values in DEFAULT_OPTIONS.items():
        for value in values:
            db.execute(
                "INSERT OR IGNORE INTO inventory_options (category, value) VALUES (?, ?)",
                (category, value),
            )
    db.commit()


def ensure_inventory_schema():
    db = get_db()
    columns = [row["name"] for row in db.execute("PRAGMA table_info(inventory)").fetchall()]
    if "category" not in columns:
        db.execute("ALTER TABLE inventory ADD COLUMN category TEXT")
        db.commit()


def ensure_admin_user():
    db = get_db()
    ensure_users_schema()
    existing = db.execute("SELECT id FROM users LIMIT 1").fetchone()
    if existing:
        return
    username = os.environ.get("ADMIN_USER", "admin")
    password = os.environ.get("ADMIN_PASS", "admin123")
    try:
        db.execute(
            "INSERT INTO users (username, password_hash, role, display_name) VALUES (?, ?, ?, ?)",
            (username, generate_password_hash(password), "admin", "Admin"),
        )
        db.commit()
    except sqlite3.OperationalError:
        ensure_users_schema()
        db.execute(
            "INSERT INTO users (username, password_hash, role, display_name) VALUES (?, ?, ?, ?)",
            (username, generate_password_hash(password), "admin", "Admin"),
        )
        db.commit()


def ensure_client_user():
    db = get_db()
    ensure_users_schema()
    existing = db.execute("SELECT id FROM users WHERE role = ? LIMIT 1", ("client",)).fetchone()
    if existing:
        return
    username = os.environ.get("CLIENT_USER", "client")
    password = os.environ.get("CLIENT_PASS", "client123")
    try:
        db.execute(
            "INSERT INTO users (username, password_hash, role, display_name) VALUES (?, ?, ?, ?)",
            (username, generate_password_hash(password), "client", "Client"),
        )
        db.commit()
    except sqlite3.OperationalError:
        ensure_users_schema()
        db.execute(
            "INSERT INTO users (username, password_hash, role, display_name) VALUES (?, ?, ?, ?)",
            (username, generate_password_hash(password), "client", "Client"),
        )
        db.commit()


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


def normalize_header(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().upper().split())


def format_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    return str(value).strip()


def row_to_dict(row):
    return {
        "id": row["id"],
        **{field: row[field] or "" for field in FIELDS},
        "created_at": row["created_at"] or "",
    }


def parse_warranty_date(value):
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.date.fromisoformat(raw)
    except ValueError:
        return None


def get_warranty_alert_groups(upcoming_days=7):
    db = get_db()
    rows = db.execute(
        """
        SELECT *
        FROM inventory
        WHERE warranty IS NOT NULL AND TRIM(warranty) <> ''
        ORDER BY warranty ASC
        """
    ).fetchall()
    today = datetime.date.today()
    upcoming_cutoff = today + datetime.timedelta(days=upcoming_days)
    overdue = []
    due_today = []
    upcoming = []
    for row in rows:
        item = row_to_dict(row)
        warranty_date = parse_warranty_date(item.get("warranty"))
        if not warranty_date:
            continue
        label = "Due today" if warranty_date == today else "Past due"
        if label == "Due today":
            message = f"Warranty of stock #{item.get('stock_no') or 'N/A'} is due today."
        else:
            message = (
                f"Warranty of stock #{item.get('stock_no') or 'N/A'} was due last "
                f"{warranty_date.strftime('%B %d, %Y')}."
            )
        alert = {
            "stock_no": item.get("stock_no") or "",
            "model": item.get("model") or "",
            "client": item.get("client") or "",
            "warranty": item.get("warranty") or "",
            "due_date": warranty_date.strftime("%B %d, %Y"),
            "due_label": label,
            "message": message,
            "due_sort": warranty_date,
        }
        if warranty_date < today:
            overdue.append(alert)
        elif warranty_date == today:
            due_today.append(alert)
        elif warranty_date <= upcoming_cutoff:
            alert["due_label"] = "Upcoming"
            alert["message"] = (
                f"Warranty of stock #{item.get('stock_no') or 'N/A'} is due on "
                f"{warranty_date.strftime('%B %d, %Y')}."
            )
            upcoming.append(alert)

    overdue.sort(key=lambda a: (a["due_sort"], a["stock_no"]))
    due_today.sort(key=lambda a: (a["due_sort"], a["stock_no"]))
    upcoming.sort(key=lambda a: (a["due_sort"], a["stock_no"]))
    return {
        "overdue": overdue,
        "due_today": due_today,
        "upcoming": upcoming,
        "today": today,
        "upcoming_cutoff": upcoming_cutoff,
    }


def get_options():
    db = get_db()
    rows = db.execute(
        "SELECT category, value, color FROM inventory_options ORDER BY category, value"
    ).fetchall()
    options = {key: [] for key in DEFAULT_OPTIONS.keys()}
    for row in rows:
        category = row["category"]
        if category in options:
            # normalize legacy status value
            val = row["value"]
            if category == 'status' and isinstance(val, str) and val.strip().upper() == 'IN STOCK':
                val = 'ON STOCK'
            options[category].append(val)
    return options


def log_audit(action, actor, details):
    db = get_db()
    db.execute(
        "INSERT INTO audit_logs (action, actor, details, created_at) VALUES (?, ?, ?, ?)",
        (action, actor, details, datetime.datetime.utcnow().isoformat()),
    )
    db.commit()


def get_current_actor():
    user_id = session.get("user_id")
    if not user_id:
        return "Unknown"
    db = get_db()
    row = db.execute(
        "SELECT COALESCE(display_name, username) AS name FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    return row["name"] if row else "Unknown"


def get_current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    db = get_db()
    row = db.execute(
        "SELECT username, role, COALESCE(display_name, username) AS display_name FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    if not row:
        return None
    name = row["display_name"] or row["username"]
    initials = "".join(part[0] for part in name.split() if part).upper()[:2]
    return {
        "name": name,
        "role": row["role"],
        "initials": initials or "?",
    }


@app.context_processor
def inject_current_user():
    groups = {"overdue": [], "due_today": [], "upcoming": []}
    today = datetime.date.today()
    try:
        groups = get_warranty_alert_groups()
        today = groups.get("today", today)
    except Exception:
        groups = {"overdue": [], "due_today": [], "upcoming": []}
    total = len(groups.get("overdue", [])) + len(groups.get("due_today", [])) + len(groups.get("upcoming", []))
    popup_messages = [a["message"] for a in groups.get("overdue", []) + groups.get("due_today", [])]
    return {
        "current_user": get_current_user(),
        "notification_count": total,
        "notification_messages": popup_messages,
        "notification_date": today.isoformat(),
    }


def get_inventory_rows(start_date=None, end_date=None, category=None, order_by="id DESC"):
    db = get_db()
    query = "SELECT * FROM inventory"
    params = []
    conditions = []

    if start_date:
        conditions.append("SUBSTR(created_at, 1, 10) >= ?")
        params.append(start_date)
    if end_date:
        conditions.append("SUBSTR(created_at, 1, 10) <= ?")
        params.append(end_date)
    if category:
        conditions.append("UPPER(TRIM(category)) = ?")
        params.append(category.upper())

    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += f" ORDER BY {order_by}"
    return db.execute(query, params).fetchall()


@app.before_request
def startup():
    global _startup_done
    if not _startup_done:
        init_db()
        ensure_users_schema()
        ensure_inventory_schema()
        ensure_admin_user()
        ensure_inventory_options()
        _startup_done = True


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            return redirect(url_for("index"))
        # on failed login, still provide the username list so the select keeps options
        users = db.execute("SELECT username FROM users ORDER BY id ASC").fetchall()
        usernames = [r['username'] for r in users]
        return render_template("login.html", error="Invalid username or password", usernames=usernames)
    # provide list of usernames for the login select (GET)
    db = get_db()
    users = db.execute("SELECT username FROM users ORDER BY id ASC").fetchall()
    usernames = [r['username'] for r in users]
    return render_template("login.html", error=None, usernames=usernames)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    db = get_db()
    # overall total
    total = db.execute("SELECT COUNT(*) FROM inventory").fetchone()[0]
    # per-category counts and last update
    def category_stats(cat):
        cat_up = db.execute("SELECT COUNT(*) FROM inventory WHERE TRIM(UPPER(category)) = ?", (cat,)).fetchone()[0]
        cat_on_stock = db.execute("SELECT COUNT(*) FROM inventory WHERE TRIM(UPPER(category)) = ? AND TRIM(UPPER(status)) = 'ON STOCK'", (cat,)).fetchone()[0]
        cat_delivered = db.execute("SELECT COUNT(*) FROM inventory WHERE TRIM(UPPER(category)) = ? AND TRIM(UPPER(status)) = 'DELIVERED'", (cat,)).fetchone()[0]
        # some DBs may not have updated_at column; fallback to created_at if missing
        try:
            last = db.execute("SELECT MAX(updated_at) FROM inventory WHERE TRIM(UPPER(category)) = ?", (cat,)).fetchone()[0]
        except sqlite3.OperationalError:
            try:
                last = db.execute("SELECT MAX(created_at) FROM inventory WHERE TRIM(UPPER(category)) = ?", (cat,)).fetchone()[0]
            except Exception:
                last = None
        return {"total": cat_up, "on_stock": cat_on_stock, "delivered": cat_delivered, "last_update": last}

    ups_stats = category_stats('UPS')
    avr_stats = category_stats('AVR')

    return render_template(
        "index.html",
        active_page="dashboard",
        total_units=total,
        ups_stats=ups_stats,
        avr_stats=avr_stats,
    )


@app.route("/inventory")
@login_required
def inventory():
    rows = get_inventory_rows()
    items = [row_to_dict(row) for row in rows]
    options = get_options()
    return render_template(
        "inventory.html",
        active_page="inventory",
        items=items,
        items_json=json.dumps(items),
        options=options,
    )


@app.route("/barcode")
@login_required
def barcode_page():
    options = get_options()
    return render_template("barcode.html", active_page="barcode", options=options)


@app.route("/clients")
@login_required
def clients():
    db = get_db()
    rows = db.execute(
        """
        SELECT
            COALESCE(NULLIF(TRIM(client), ''), 'Unspecified') AS client,
            COUNT(*) AS total_units,
            MAX(date_acquired) AS last_acquired,
            MAX(date_installed) AS last_installed
        FROM inventory
        GROUP BY COALESCE(NULLIF(TRIM(client), ''), 'Unspecified')
        ORDER BY client ASC
        """
    ).fetchall()

    clients_data = [
        {
            "client": row["client"],
            "total_units": row["total_units"],
            "last_acquired": row["last_acquired"] or "-",
            "last_installed": row["last_installed"] or "-",
        }
        for row in rows
    ]

    return render_template(
        "clients.html",
        active_page="clients",
        clients=clients_data,
        total_clients=len(clients_data),
    )


@app.route("/installations")
@login_required
def installations():
    db = get_db()
    rows = db.execute(
        """
        SELECT *
        FROM inventory
        WHERE date_installed IS NOT NULL AND TRIM(date_installed) <> ''
        ORDER BY date_installed DESC
        """
    ).fetchall()
    items = [row_to_dict(row) for row in rows]
    return render_template(
        "installations.html",
        active_page="installations",
        items=items,
        total_installations=len(items),
    )


@app.route("/warranty")
@login_required
def warranty():
    db = get_db()
    rows = db.execute(
        """
        SELECT *
        FROM inventory
        WHERE warranty IS NOT NULL AND TRIM(warranty) <> ''
        ORDER BY warranty DESC
        """
    ).fetchall()
    items = [row_to_dict(row) for row in rows]
    return render_template(
        "warranty.html",
        active_page="warranty",
        items=items,
        total_warranty=len(items),
    )


@app.route("/notifications")
@login_required
def notifications():
    groups = get_warranty_alert_groups()
    today = groups.get("today", datetime.date.today())
    overdue = groups.get("overdue", [])
    due_today = groups.get("due_today", [])
    upcoming = groups.get("upcoming", [])
    total_alerts = len(overdue) + len(due_today) + len(upcoming)
    return render_template(
        "notifications.html",
        active_page="notifications",
        overdue=overdue,
        due_today=due_today,
        upcoming=upcoming,
        total_alerts=total_alerts,
        due_today_count=len(due_today),
        overdue_count=len(overdue),
        upcoming_count=len(upcoming),
        today=today.strftime("%B %d, %Y"),
    )


@app.route("/reports")
@login_required
def reports():
    db = get_db()
    total_units = db.execute("SELECT COUNT(*) FROM inventory").fetchone()[0]
    status_rows = db.execute(
        """
        SELECT COALESCE(status, '') AS status, COUNT(*) AS count
        FROM inventory
        GROUP BY status
        """
    ).fetchall()
    client_rows = db.execute(
        """
        SELECT COALESCE(NULLIF(TRIM(client), ''), 'Unspecified') AS client, COUNT(*) AS count
        FROM inventory
        GROUP BY COALESCE(NULLIF(TRIM(client), ''), 'Unspecified')
        ORDER BY count DESC
        """
    ).fetchall()
    model_rows = db.execute(
        """
        SELECT COALESCE(NULLIF(TRIM(model), ''), 'Unspecified') AS model, COUNT(*) AS count
        FROM inventory
        GROUP BY COALESCE(NULLIF(TRIM(model), ''), 'Unspecified')
        ORDER BY count DESC
        """
    ).fetchall()
    rating_rows = db.execute(
        """
        SELECT COALESCE(NULLIF(TRIM(rating), ''), 'Unspecified') AS rating, COUNT(*) AS count
        FROM inventory
        GROUP BY COALESCE(NULLIF(TRIM(rating), ''), 'Unspecified')
        ORDER BY count DESC
        """
    ).fetchall()
    system_rows = db.execute(
        """
        SELECT COALESCE(NULLIF(TRIM(system), ''), 'Unspecified') AS system, COUNT(*) AS count
        FROM inventory
        GROUP BY COALESCE(NULLIF(TRIM(system), ''), 'Unspecified')
        ORDER BY count DESC
        """
    ).fetchall()
    month_rows = db.execute(
        """
        SELECT
            COALESCE(NULLIF(SUBSTR(TRIM(date_acquired), 1, 7), ''), 'Unspecified') AS month,
            COUNT(*) AS count
        FROM inventory
        GROUP BY COALESCE(NULLIF(SUBSTR(TRIM(date_acquired), 1, 7), ''), 'Unspecified')
        ORDER BY month ASC
        """
    ).fetchall()

    status_counts = [
        {"status": row["status"] or "Unspecified", "count": row["count"]}
        for row in status_rows
    ]
    top_clients = [
        {"client": row["client"], "count": row["count"]} for row in client_rows
    ]
    model_counts = [{"model": row["model"], "count": row["count"]} for row in model_rows]
    rating_counts = [
        {"rating": row["rating"], "count": row["count"]} for row in rating_rows
    ]
    system_counts = [
        {"system": row["system"], "count": row["count"]} for row in system_rows
    ]
    month_counts = [{"month": row["month"], "count": row["count"]} for row in month_rows]

    options = get_options()
    # category counts
    category_rows = db.execute(
        "SELECT COALESCE(NULLIF(TRIM(category),''),'Unspecified') AS category, COUNT(*) AS count FROM inventory GROUP BY COALESCE(NULLIF(TRIM(category),''),'Unspecified')"
    ).fetchall()
    category_counts = [{"category": r["category"], "count": r["count"]} for r in category_rows]
    return render_template(
        "reports.html",
        active_page="reports",
        total_units=total_units,
        status_counts=status_counts,
        top_clients=top_clients,
        model_counts=model_counts,
        rating_counts=rating_counts,
        system_counts=system_counts,
        month_counts=month_counts,
        option_categories=CATEGORY_META,
        options=options,
        category_counts=category_counts,
    )


@app.route("/options")
@login_required
def options_page():
    db = get_db()
    rows = db.execute(
        "SELECT id, category, value, color FROM inventory_options ORDER BY category, value"
    ).fetchall()
    options = {meta["key"]: [] for meta in CATEGORY_META}
    for row in rows:
        category = row["category"]
        if category in options:
            try:
                color = row["color"]
            except Exception:
                color = None
            options[category].append({"id": row["id"], "value": row["value"], "color": color})

    return render_template(
        "options.html",
        active_page="options",
        categories=CATEGORY_META,
        options=options,
    )


@app.route("/accounts", methods=["GET", "POST"])
@login_required
def accounts():
    db = get_db()
    error = None

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        role = request.form.get("role", "staff").strip() or "staff"

        if not username or not password:
            error = "Username and password are required."
        else:
            try:
                db.execute(
                    "INSERT INTO users (username, password_hash, role, display_name) VALUES (?, ?, ?, ?)",
                    (username, generate_password_hash(password), role, name),
                )
                db.commit()
                log_audit(
                    "Create account",
                    get_current_actor(),
                    f"Created user '{username}' with role '{role}'.",
                )
            except sqlite3.IntegrityError:
                error = "Username already exists."

    rows = db.execute(
        "SELECT id, username, role, COALESCE(display_name, '') AS display_name FROM users ORDER BY id ASC"
    ).fetchall()
    users = [
        {
            "id": row["id"],
            "username": row["username"],
            "role": row["role"],
            "display_name": row["display_name"] or row["username"],
        }
        for row in rows
    ]

    logs = db.execute(
        "SELECT action, actor, details, created_at FROM audit_logs ORDER BY id DESC LIMIT 100"
    ).fetchall()
    audit_logs = [
        {
            "action": row["action"],
            "actor": row["actor"],
            "details": row["details"],
            "created_at": row["created_at"],
        }
        for row in logs
    ]

    return render_template(
        "accounts.html",
        active_page="accounts",
        users=users,
        error=error,
        audit_logs=audit_logs,
    )


@app.route('/api/audit', methods=['GET'])
@login_required
def api_audit_list():
    db = get_db()
    rows = db.execute("SELECT id, action, actor, details, created_at FROM audit_logs ORDER BY id DESC LIMIT 200").fetchall()
    return jsonify({"logs": [{"id": r["id"], "action": r["action"], "actor": r["actor"], "details": r["details"], "created_at": r["created_at"]} for r in rows]})


@app.route('/api/undo/<int:log_id>', methods=['POST'])
@login_required
def api_undo(log_id):
    db = get_db()
    row = db.execute("SELECT id, action, actor, details FROM audit_logs WHERE id = ?", (log_id,)).fetchone()
    if not row:
        return jsonify({"error": "log not found"}), 404
    action = row["action"] or ""
    details = row["details"] or ""
    # Support undo for options add/delete only
    try:
        if action == 'Options' and details.startswith('Added option'):
            # details: Added option {value} to {category}
            parts = details.split('Added option', 1)[1].strip()
            # naive parse
            if ' to ' in parts:
                val, cat = parts.split(' to ', 1)
                val = val.strip()
                cat = cat.strip()
                db.execute("DELETE FROM inventory_options WHERE category = ? AND value = ?", (cat, val))
                db.commit()
                log_audit('Undo', get_current_actor(), f"Undid add option {val} from {cat}")
                return jsonify({"status": "ok"})
        if action == 'Options' and details.startswith('Deleted option'):
            # details: Deleted option {value} from {category}
            parts = details.split('Deleted option', 1)[1].strip()
            if ' from ' in parts:
                val, cat = parts.split(' from ', 1)
                val = val.strip()
                cat = cat.strip()
                db.execute("INSERT OR IGNORE INTO inventory_options (category, value) VALUES (?, ?)", (cat, val))
                db.commit()
                log_audit('Undo', get_current_actor(), f"Undid delete option {val} in {cat}")
                return jsonify({"status": "ok"})
    except Exception:
        db.rollback()
        return jsonify({"error": "undo failed"}), 500
    return jsonify({"error": "undo not supported for this log"}), 400


@app.route("/accounts/delete", methods=["POST"])
@login_required
def accounts_delete():
    user_id = request.form.get("id", "").strip()
    if user_id:
        db = get_db()
        user_row = db.execute(
            "SELECT username, role FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
        db.execute("DELETE FROM users WHERE id = ?", (user_id,))
        db.commit()
        if user_row:
            log_audit(
                "Delete account",
                get_current_actor(),
                f"Deleted user '{user_row['username']}' with role '{user_row['role']}'.",
            )
    return redirect(url_for("accounts"))


@app.route("/accounts/update", methods=["POST"])
@login_required
def accounts_update():
    user_id = request.form.get("id", "").strip()
    name = request.form.get("name", "").strip()
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    role = request.form.get("role", "staff").strip() or "staff"

    if not user_id or not username:
        return redirect(url_for("accounts"))

    db = get_db()
    before = db.execute(
        "SELECT username, role FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()

    try:
        if password:
            db.execute(
                """
                UPDATE users
                SET username = ?, role = ?, display_name = ?, password_hash = ?
                WHERE id = ?
                """,
                (username, role, name, generate_password_hash(password), user_id),
            )
        else:
            db.execute(
                """
                UPDATE users
                SET username = ?, role = ?, display_name = ?
                WHERE id = ?
                """,
                (username, role, name, user_id),
            )
        db.commit()
        if before:
            log_audit(
                "Update account",
                get_current_actor(),
                f"Updated user '{before['username']}' to '{username}', role '{role}'.",
            )
    except sqlite3.IntegrityError:
        pass

    return redirect(url_for("accounts"))


@app.route("/options/add", methods=["POST"])
@login_required
def options_add():
    category = request.form.get("category", "").strip()
    value = request.form.get("value", "").strip()
    color = request.form.get("color", "").strip() or None
    # only accept color for category options (UPS/AVR)
    if category != 'category':
        color = None
    if category in DEFAULT_OPTIONS and value:
        db = get_db()
        try:
            db.execute(
                "INSERT OR IGNORE INTO inventory_options (category, value, color) VALUES (?, ?, ?)",
                (category, value, color),
            )
            db.commit()
            log_audit("Options", get_current_actor(), f"Added option {value} to {category}")
        except Exception:
            db.rollback()
    return redirect(url_for("options_page"))


@app.route("/options/update", methods=["POST"])
@login_required
def options_update():
    option_id = request.form.get("id", "").strip()
    value = request.form.get("value", "").strip()
    color = request.form.get("color", "").strip() or None
    if option_id and value:
        db = get_db()
        before = db.execute("SELECT category, value FROM inventory_options WHERE id = ?", (option_id,)).fetchone()
        # only store color when this option belongs to category
        cat = before['category'] if before else None
        if cat == 'category':
            db.execute(
                "UPDATE inventory_options SET value = ?, color = ? WHERE id = ?",
                (value, color, option_id),
            )
        else:
            db.execute(
                "UPDATE inventory_options SET value = ? WHERE id = ?",
                (value, option_id),
            )
        db.commit()
        if before:
            log_audit("Options", get_current_actor(), f"Updated option {before['value']} -> {value} in {before['category']}")
    return redirect(url_for("options_page"))


@app.route("/options/delete", methods=["POST"])
@login_required
def options_delete():
    option_id = request.form.get("id", "").strip()
    if option_id:
        db = get_db()
        row = db.execute("SELECT category, value FROM inventory_options WHERE id = ?", (option_id,)).fetchone()
        if row:
            # prevent deleting an option that is in use
            used = db.execute(
                "SELECT 1 FROM inventory WHERE TRIM(COALESCE(%s,'')) = TRIM(?) LIMIT 1" % (row["category"]),
                (row["value"],),
            ).fetchone()
            if used:
                # do not delete if still referenced
                return redirect(url_for("options_page"))
            db.execute("DELETE FROM inventory_options WHERE id = ?", (option_id,))
            db.commit()
            log_audit("Options", get_current_actor(), f"Deleted option {row['value']} from {row['category']}")
    return redirect(url_for("options_page"))


@app.route("/api/items", methods=["GET", "POST"])
@login_required
def items():
    db = get_db()
    if request.method == "POST":
        payload = request.get_json(silent=True) or {}
        serial = payload.get("serial_number", "").strip()
        if serial:
            existing = db.execute(
                "SELECT id FROM inventory WHERE TRIM(serial_number) = ? LIMIT 1",
                (serial,),
            ).fetchone()
            if existing:
                return jsonify({"error": "Serial Number already exists."}), 400

        try:
            db.execute("BEGIN IMMEDIATE")
        except Exception:
            pass

        try:
            row = db.execute(
                "SELECT MAX(CAST(stock_no AS INTEGER)) AS max_stock FROM inventory"
            ).fetchone()
            max_stock = row["max_stock"] if row and row["max_stock"] is not None else 0
            next_stock = int(max_stock) + 1
        except Exception:
            next_stock = 1

        values = [payload.get(field, "").strip() for field in FIELDS]
        # server-side validation: ensure provided category and system exist in options
        supplied_category = values[ FIELDS.index('category') ] if 'category' in FIELDS else ''
        supplied_system = values[ FIELDS.index('system') ] if 'system' in FIELDS else ''
        db = get_db()
        if supplied_category:
            exists = db.execute(
                "SELECT 1 FROM inventory_options WHERE category='category' AND TRIM(UPPER(value)) = TRIM(UPPER(?)) LIMIT 1",
                (supplied_category,),
            ).fetchone()
            if not exists:
                return jsonify({"error": "Invalid category"}), 400
        if supplied_system:
            exists = db.execute(
                "SELECT 1 FROM inventory_options WHERE category='brand' AND TRIM(UPPER(value)) = TRIM(UPPER(?)) LIMIT 1",
                (supplied_system,),
            ).fetchone()
            if not exists:
                return jsonify({"error": "Invalid brand/system"}), 400
        values[0] = str(next_stock)
        db.execute(
            """
            INSERT INTO inventory (
                stock_no, code, category, system, model, rating, status, serial_number, client,
                date_acquired, date_installed, dr_no, si_no, po, value_vat_ex, warranty,
                terms, remarks, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [*values, datetime.datetime.utcnow().isoformat()],
        )
        db.commit()
        new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        row = db.execute("SELECT * FROM inventory WHERE id = ?", (new_id,)).fetchone()
        actor = get_current_actor()
        details = f"{actor} added new item in inventory: {row['model'] or 'Unspecified'} / {row['serial_number'] or 'N/A'}."
        log_audit("Add inventory", actor, details)
        return jsonify({"item": row_to_dict(row)})

    rows = db.execute("SELECT * FROM inventory ORDER BY id DESC").fetchall()
    return jsonify({"items": [row_to_dict(row) for row in rows]})


@app.route("/api/items/<int:item_id>", methods=["PUT", "DELETE"])
@login_required
def item_detail(item_id):
    db = get_db()
    if request.method == "DELETE":
        row = db.execute("SELECT * FROM inventory WHERE id = ?", (item_id,)).fetchone()
        db.execute("DELETE FROM inventory WHERE id = ?", (item_id,))
        db.commit()
        if row:
            actor = get_current_actor()
            details = f"{actor} removed item from inventory: {row['model'] or 'Unspecified'} / {row['serial_number'] or 'N/A'}."
            log_audit("Remove inventory", actor, details)
        return jsonify({"status": "ok"})

    payload = request.get_json(silent=True) or {}
    serial = payload.get("serial_number", "").strip()
    if serial:
        existing = db.execute(
            "SELECT id FROM inventory WHERE TRIM(serial_number) = ? AND id != ? LIMIT 1",
            (serial, item_id),
        ).fetchone()
        if existing:
            return jsonify({"error": "Serial Number already exists."}), 400
    values = [payload.get(field, "").strip() for field in FIELDS]
    # validate category and system on update as well
    supplied_category = values[ FIELDS.index('category') ] if 'category' in FIELDS else ''
    supplied_system = values[ FIELDS.index('system') ] if 'system' in FIELDS else ''
    db = get_db()
    if supplied_category:
        exists = db.execute(
            "SELECT 1 FROM inventory_options WHERE category='category' AND TRIM(UPPER(value)) = TRIM(UPPER(?)) LIMIT 1",
            (supplied_category,),
        ).fetchone()
        if not exists:
            return jsonify({"error": "Invalid category"}), 400
    if supplied_system:
        exists = db.execute(
            "SELECT 1 FROM inventory_options WHERE category='brand' AND TRIM(UPPER(value)) = TRIM(UPPER(?)) LIMIT 1",
            (supplied_system,),
        ).fetchone()
        if not exists:
            return jsonify({"error": "Invalid brand/system"}), 400
    db.execute(
        """
        UPDATE inventory SET
            stock_no = ?, code = ?, category = ?, system = ?, model = ?, rating = ?, status = ?,
            serial_number = ?, client = ?, date_acquired = ?, date_installed = ?,
            dr_no = ?, si_no = ?, po = ?, value_vat_ex = ?, warranty = ?, terms = ?,
            remarks = ?
        WHERE id = ?
        """,
        [*values, item_id],
    )
    db.commit()
    row = db.execute("SELECT * FROM inventory WHERE id = ?", (item_id,)).fetchone()
    actor = get_current_actor()
    details = f"{actor} updated inventory item: {row['model'] or 'Unspecified'} / {row['serial_number'] or 'N/A'}."
    log_audit("Update inventory", actor, details)
    return jsonify({"item": row_to_dict(row)})


@app.route("/api/import", methods=["POST"])
@login_required
def import_excel():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    try:
        from io import BytesIO

        file_bytes = file.read()
        stream = BytesIO(file_bytes)
        wb = load_workbook(stream, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("No active worksheet found in uploaded workbook")
    except Exception as e:
        app.logger.exception("Failed to load uploaded Excel file")
        return jsonify({"error": f"Failed to read Excel file: {str(e)}"}), 400

    header_cells = [normalize_header(cell.value) for cell in ws[1]]
    mapping = {}
    for idx, header in enumerate(header_cells):
        if header in HEADER_MAP:
            mapping[idx] = HEADER_MAP[header]

    # optional category override from form (e.g., UPS or AVR)
    override_category = (request.form.get('category') or '').strip()

    if not mapping:
        return jsonify({"error": "No matching headers found"}), 400

    db = get_db()
    created = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2):
        record = {field: "" for field in FIELDS}
        for idx, cell in enumerate(row):
            if idx in mapping:
                record[mapping[idx]] = format_cell(cell.value)

        # if override_category provided, set the category for every imported record
        if override_category:
            record['category'] = override_category
        if all(not record[field] for field in FIELDS):
            skipped += 1
            continue

        values = [record[field] for field in FIELDS]
        db.execute(
            """
            INSERT INTO inventory (
                stock_no, code, category, system, model, rating, status, serial_number, client,
                date_acquired, date_installed, dr_no, si_no, po, value_vat_ex, warranty,
                terms, remarks, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [*values, datetime.datetime.utcnow().isoformat()],
        )
        created += 1

    db.commit()
    return jsonify({"created": created, "skipped": skipped})


def create_excel_sheet(ws, headers, rows, tab_color):
    ws.sheet_properties.tabColor = f"FF{tab_color}"
    ws.append(headers)

    for row in rows:
        ws.append(row)

    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ws.columns:
        max_length = 0
        for cell in col:
            value = str(cell.value or "")
            max_length = max(max_length, len(value))
        adjusted_width = min(max(max_length + 4, 14), 80)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    status_index = None
    if "STATUS" in headers:
        status_index = headers.index("STATUS")

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if isinstance(cell.value, str):
                cell.number_format = '@'
            if status_index is not None and idx == status_index and isinstance(cell.value, str):
                value = cell.value.strip().upper()
                if value == "ON STOCK":
                    cell.fill = green_fill
                elif value == "DELIVERED":
                    cell.fill = red_fill

    ws.freeze_panes = "A2"


@app.route("/api/export")
@login_required
def export_excel():
    try:
        start_date = request.args.get("start") or None
        end_date = request.args.get("end") or None
        category = request.args.get("category") or None
        rows = get_inventory_rows(start_date, end_date, category=category, order_by="id ASC")

        wb = Workbook()
        inventory_ws = wb.active
        if inventory_ws is None:
            raise RuntimeError("Failed to create inventory worksheet")
        inventory_ws.title = "Inventory"

        inventory_headers = [
            "STOCK #",
            "CODE",
            "SYSTEM",
            "MODEL",
            "RATING",
            "STATUS",
            "SERIAL NUMBER",
            "CLIENT",
            "DATE ACQUIRED",
            "DATE INSTALLED",
            "DR NO.",
            "SI NO.",
            "PO",
            "VALUE VAT EX",
            "WARRANTY",
            "TERMS",
            "REMARKS",
        ]
        inventory_rows = [[row[field] or "" for field in FIELDS] for row in rows]

        create_excel_sheet(inventory_ws, inventory_headers, inventory_rows, "92D050")

        # prepare response
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="inventory.xlsx",
        )
    except Exception as e:
        app.logger.exception("Export failed")
        return jsonify({"error": f"Export failed: {str(e)}"}), 500


@app.route("/export/pdf")
@login_required
def export_pdf():
    start_date = request.args.get("start") or None
    end_date = request.args.get("end") or None
    rows = get_inventory_rows(start_date, end_date, order_by="id ASC")
    items = [row_to_dict(row) for row in rows]
    return render_template(
        "export_pdf.html",
        items=items,
        start_date=start_date,
        end_date=end_date,
    )

# PDF export removed. Route deprecated.


@app.route("/api/summary")
@login_required
def summary():
    db = get_db()
    total_count = db.execute("SELECT COUNT(*) FROM inventory").fetchone()[0]
    status_rows = db.execute(
        "SELECT COALESCE(status, '') AS status, COUNT(*) AS count FROM inventory GROUP BY status"
    ).fetchall()
    model_rows = db.execute(
        """
        SELECT COALESCE(model, '') AS model, COUNT(*) AS count
        FROM inventory
        WHERE model IS NOT NULL AND TRIM(model) <> ''
        GROUP BY model
        ORDER BY count DESC
        LIMIT 5
        """
    ).fetchall()

    # build per-category stats for UPS and AVR
    def cat_stats(cat):
        total = db.execute("SELECT COUNT(*) FROM inventory WHERE TRIM(UPPER(category)) = ?", (cat,)).fetchone()[0]
        on_stock = db.execute("SELECT COUNT(*) FROM inventory WHERE TRIM(UPPER(category)) = ? AND TRIM(UPPER(status)) = 'ON STOCK'", (cat,)).fetchone()[0]
        delivered = db.execute("SELECT COUNT(*) FROM inventory WHERE TRIM(UPPER(category)) = ? AND TRIM(UPPER(status)) = 'DELIVERED'", (cat,)).fetchone()[0]
        return {"total": total, "on_stock": on_stock, "delivered": delivered}

    category_counts = {"UPS": cat_stats("UPS"), "AVR": cat_stats("AVR")}

    return jsonify(
        {
            "total_count": total_count,
            "status_counts": [
                {"status": row["status"] or "Unspecified", "count": row["count"]}
                for row in status_rows
            ],
            "model_counts": [
                {"model": row["model"], "count": row["count"]} for row in model_rows
            ],
            "category_counts": category_counts,
        }
    )


def open_default_browser():
    try:
        webbrowser.open("http://127.0.0.1:5000/login")
    except Exception:
        pass

if __name__ == "__main__":
    with app.app_context():
        init_db()
        ensure_users_schema()
        ensure_inventory_schema()
        ensure_admin_user()
        # Do not auto-create client users

    threading.Timer(1.5, open_default_browser).start()
    app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=True)


@app.errorhandler(500)
def handle_500(e):
    tb = traceback.format_exc()
    app.logger.error("Internal server error: %s", tb)
    # return JSON in API calls, otherwise render simple page
    if request.path.startswith('/api/'):
        return jsonify({"error": "Internal server error", "details": str(e)}), 500
    return render_template('500.html', error=str(e)), 500
